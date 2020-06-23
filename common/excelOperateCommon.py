#!/usr/bin/env python
# -*- coding: utf-8 -*-

# ------
# @Description ：读取Excel数据
# @File        : constant.py
# @Date        : 2020-05-06
# @Author      : MiaWang
# @copyright   : Winning HealthCare
# ------

from openpyxl import load_workbook
from common.logger import Logger


class ExcelOperateCommonUtil:
    def __init__(self, file_name, sheet_name):
        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_excel(self):
        Logger.debug("TestDataReaderCommonUtil.read_excel file name is : " + self.file_name,
                     self.__class__.__name__)
        excel = load_workbook(self.file_name)
        return excel

    def read_excel_sheet(self):
        Logger.debug("TestDataReaderCommonUtil.read_excel_sheet file name is : {0}, sheet_name is : {1}"
                     .format(self.file_name, self.sheet_name), self.__class__.__name__)
        excel = load_workbook(self.file_name)
        sheet = excel[self.sheet_name]
        return sheet

    def get_row_count(self):
        """获取excel的行数"""
        Logger.debug("TestDataReaderCommonUtil.get_row_count 获取excel的行数 file name is : " + self.file_name,
                     self.__class__.__name__)
        return self.read_excel_sheet().max_row

    def get_cel_value_by_row_col(self, row, col):
        """
        获取excel中单元格内容
        :param row:
        :param col:
        :return:
        """
        Logger.debug("TestDataReaderCommonUtil.get_cel_value_by_row_col 获取excel中单元格内容 "
                     "file name is : {0}, row:{1}, col:{2}".format(self.file_name, row, col),
                     self.__class__.__name__)
        cell_type = type(self.read_excel_sheet().cell(row=row, column=col).value)
        cell_value = self.read_excel_sheet().cell(row=row, column=col).value
        # 整数类型去除小数点，如dept_code，读取出来是3203.0格式的，去除小数点后为3203格式
        if cell_type == float and cell_value % 1 == 0:
            cell_value = int(cell_value)
        return cell_value

    def write_result(self, row, col, content):
        """
        将实际结果写到测试数据excel中
        :param row: 
        :param col: 
        :param content: 
        :return:
        """
        excel = self.read_excel()
        sheet = excel[self.sheet_name]
        sheet.cell(row, col).value = content
        excel.save(self.file_name)
